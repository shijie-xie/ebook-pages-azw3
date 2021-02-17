import os
import re
f = open("test.txt")                  
line = f.readline()   
pages = 1240   #first page of index
list1 = []
while line:   
    #print(line) 
    if line.replace(" ",'') == "\n":
        pages = pages +1 #I copied the pdf to a txt file. if there is a blank line. It should be the next page.
                    
    #print(line, end = '')　     
    num = re.findall('\d+',line)
    #print(num,line)   
    for i in num:
        if int(i)<1000:
            list1.append((int(i),line,pages))
    line = f.readline()   
list1.sort()
#print(list1)



   
f.close()

#write to a xls file
import openpyxl
wb=openpyxl.load_workbook('test.xlsx')
ws=wb.worksheets[0]
for t in list1:
    ws.cell(t[0],1,str(t[0]))
    ws.cell(t[0],2,t[1])
    ws.cell(t[0],3,t[2])
    print(t[0])


wb.save('test.xlsx')
